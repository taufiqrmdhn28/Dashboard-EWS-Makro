import os
import time
import shutil
import numpy as np
import pandas as pd

from .data import read_data, build_calendar, build_pseudo_real_time_vintages
from .prediction import (
    get_current_quarter,
    update_predictions,
    compute_annual_nowcast,
    nowcast_current_output
)
from .estimation import (
    fit_dfm_em,
    mixed_frequency_restrictions,
    idiosyncratic_law_of_motion,
    evaluate_em_convergence
)
from .benchmark import (
    ar_realtime_gdp,
    update_benchmark_predictions,
    compute_forecast_errors
)

def run_nowcasting_pipeline(
    data_file,
    hist_eval=None,
    start_est=None,
    start_eval=None,
    last_eval=None,
    ser_news=None,
    freq_estimation=None,
    growth_rate=None,
    idiosyncratic=None,
    p_ar=None,
    native=None,
    output_dir=None,
    debug_dir=None,
    threshold=None,
    threshold_direction=None,
    prob_distribution=None,
    config_path=None,
    estimator=None,
    robust_scaling=None,
    winsorization=None,
    winsorization_k=None,
    covariance_regularization=None,
    ridge_lambda=None,
    mcmc_draws=None,
    mcmc_burnin=None,
    prob_bins=None,
    generate_pptx=True
):
    from .config import load_config
    
    # 1. Load config from file/env
    config = load_config(config_path)
    
    # 2. Resolve parameter overrides
    if hist_eval is None:
        hist_eval = config.get("hist_eval", 1)
    if start_est is None:
        start_est = config.get("start_est", None)
    if start_eval is None:
        start_eval = config.get("start_eval", None)
    if last_eval is None:
        last_eval = config.get("last_eval", None)
    if ser_news is None:
        ser_news = config.get("ser_news", "RGDP_growth")
    if freq_estimation is None:
        freq_estimation = config.get("freq_estimation", "quarterly")
    if growth_rate is None:
        growth_rate = config.get("growth_rate", "yoy")
    if idiosyncratic is None:
        idiosyncratic = config.get("idiosyncratic", "Autoregressive")
    if p_ar is None:
        p_ar = config.get("p_ar", 2)
    if native is None:
        native = config.get("native", False)
    if output_dir is None:
        output_dir = config.get("output_dir", "temp/python")
    if debug_dir is None:
        debug_dir = config.get("debug_dir", None)
    if threshold is None:
        threshold = config.get("threshold", None)
    if threshold_direction is None:
        threshold_direction = config.get("threshold_direction", "lower")
    if prob_distribution is None:
        prob_distribution = config.get("prob_distribution", "gaussian")

    if estimator is None:
        estimator = config.get("estimator", "EM")
    if robust_scaling is None:
        robust_scaling = config.get("robust_scaling", False)
    if winsorization is None:
        winsorization = config.get("winsorization", False)
    if winsorization_k is None:
        winsorization_k = config.get("winsorization_k", 4.0)
    if covariance_regularization is None:
        covariance_regularization = config.get("covariance_regularization", False)
    if ridge_lambda is None:
        ridge_lambda = config.get("ridge_lambda", 1e-4)
    if mcmc_draws is None:
        mcmc_draws = config.get("mcmc_draws", 100)
    if mcmc_burnin is None:
        mcmc_burnin = config.get("mcmc_burnin", 50)

    seq_threshold = config.get("seq_threshold", 0.0)
    seq_direction = config.get("seq_direction", "lower")
    seq_n_consecutive = config.get("seq_n_consecutive", 2)
    seq_horizon = config.get("seq_horizon", 4)
    seq_k_offset = config.get("seq_k_offset", 0)

    if debug_dir:
        if os.path.exists(debug_dir):
            try:
                shutil.rmtree(debug_dir)
            except OSError:
                pass
        os.makedirs(debug_dir, exist_ok=True)
        os.environ["DFM_NOWCAST_DEBUG_DIR"] = debug_dir
        # Reset debug flag in kalman run_kf to allow initial exports
        from .kalman import reset_kf_run_flag
        reset_kf_run_flag()

    print(f"Reading and transforming data from: {data_file}")
    X_Last, Series, ReleaseDay, SeriesQ, Date, nQ = read_data(data_file, start_est)

    T, N = X_Last.shape
    nM = N - nQ

    print(f"Building calendar and vintages...")
    name_vintage, PMV, MM, R_cal = build_calendar(
        data_file, start_eval, last_eval, nQ, hist_eval=(hist_eval == 1)
    )
    DD, DDname, V, DatesV = build_pseudo_real_time_vintages(
        X_Last, Date, name_vintage, MM, R_cal, PMV
    )

    # Resolve DFM parameters from config
    p_val = config.get('p', 2)
    max_iter_val = config.get('max_iter', 300)
    thresh_val = config.get('thresh', 1e-3)
    
    r_val = config.get('r', 2)
    if isinstance(r_val, (int, float)):
        r_arr = int(r_val) * np.ones(1, dtype=int)
    else:
        r_arr = np.array(r_val, dtype=int).flatten()
        
    dyn_val = config.get('dyn', 2)
    if isinstance(dyn_val, (int, float)):
        dyn_arr = int(dyn_val) * np.ones(1, dtype=int)
    else:
        dyn_arr = np.array(dyn_val, dtype=int).flatten()

    # Initialize model estimation parameter structure P
    P = {
        'p': p_val,
        'blocks': np.ones((N, 1)),
        'r': r_arr,
        'dyn': dyn_arr,
        'nQ': nQ,
        'nM': nM,
        'Series': Series,
        'max_iter': max_iter_val,
        'thresh': thresh_val,
        'estimator': estimator,
        'robust_scaling': robust_scaling,
        'winsorization': winsorization,
        'winsorization_k': winsorization_k,
        'covariance_regularization': covariance_regularization,
        'ridge_lambda': ridge_lambda,
        'mcmc_draws': mcmc_draws,
        'mcmc_burnin': mcmc_burnin
    }
    Group = np.array(['x'] * N).reshape(N, 1)

    try:
        iSer = np.array([Series.index(ser_news)])
    except ValueError:
        print(f"Warning: '{ser_news}' not found in Series list.")
        iSer = np.array([])

    if len(iSer) == 0:
        raise ValueError(f"Target series '{ser_news}' not found in data.")

    i_ser_idx = int(iSer[0])

    t1_restr, t2_restr = mixed_frequency_restrictions(growth_rate)
    P['Rconstr'] = t1_restr
    P['q'] = t2_restr

    t1_idio = idiosyncratic_law_of_motion(idiosyncratic, nM, nQ)
    P['i_idio'] = t1_idio

    n_vintage = len(name_vintage)
    Nowcast = np.zeros((n_vintage, 1))
    Forecast = np.zeros((n_vintage, 1))
    Backcast = np.zeros((n_vintage, 1))
    Forecast2S = np.zeros((n_vintage, 1))
    Forecast3S = np.zeros((n_vintage, 1))

    Actual = np.zeros((n_vintage, 5))
    Reference = np.zeros((n_vintage, 5))

    AnnualPrediction = np.zeros((n_vintage, 1))
    AnnualGDP = np.zeros((n_vintage, 1))

    ReleaseTime = np.empty((n_vintage, 1), dtype=object)

    RW = np.zeros((n_vintage, 1))
    AR_Back = np.zeros((n_vintage, 1))
    AR_Now = np.zeros((n_vintage, 1))
    AR_Fore = np.zeros((n_vintage, 1))

    Std_Backcast = np.zeros((n_vintage, 1))
    Std_Nowcast = np.zeros((n_vintage, 1))
    Std_Forecast = np.zeros((n_vintage, 1))
    Std_Forecast2S = np.zeros((n_vintage, 1))
    Std_Forecast3S = np.zeros((n_vintage, 1))
    Std_Annual = np.zeros((n_vintage, 1))
    Prob_Seq = np.zeros((n_vintage, 1))

    if threshold is not None:
        Prob_Backcast = np.zeros((n_vintage, 1))
        Prob_Nowcast = np.zeros((n_vintage, 1))
        Prob_Forecast = np.zeros((n_vintage, 1))
        Prob_Forecast2S = np.zeros((n_vintage, 1))
        Prob_Forecast3S = np.zeros((n_vintage, 1))
        Prob_Annual = np.zeros((n_vintage, 1))

    fe_m = [[[] for _ in range(3)] for _ in range(P['nM'] + P['nQ'])]
    current_output = [[[] for _ in range(5)] for _ in range(n_vintage)]

    print("Starting Nowcasting Loop...")
    start_time = time.time()

    if hist_eval == 0:
        vintages_to_run = [n_vintage - 1]
        print(f"Monthly Monitoring Mode: Running only latest vintage (index {n_vintage - 1}).")
    else:
        vintages_to_run = range(n_vintage)
        print(f"Historical Evaluation Mode: Running all {n_vintage} vintages.")

    # Variable to cache estimated model
    R_new = None

    for t in vintages_to_run:
        ReleaseTime[t, 0] = DDname[t]
        
        # Convert Excel serial date to Python datetime to extract [Y, M, D, H, M, S]
        dt = pd.to_datetime(name_vintage[t] - 693960, unit='D', origin='1899-12-30')
        CurrDates = np.array([[dt.year, dt.month, dt.day, dt.hour, dt.minute, dt.second]], dtype=float)
        X_new = DD[t].copy()
        
        # State management for previous vintage
        if t > vintages_to_run[0]:
            prev_dt = pd.to_datetime(name_vintage[t-1] - 693960, unit='D', origin='1899-12-30')
            PrevDates = np.array([[prev_dt.year, prev_dt.month, prev_dt.day, prev_dt.hour, prev_dt.minute, prev_dt.second]], dtype=float)
            X_old = DD[t-1].copy()
        else:
            PrevDates = CurrDates
            X_old = X_new.copy()
        nobs = X_new.shape[0]
        
        # Identify current quarter end-month
        iQ = get_current_quarter(name_vintage, t, DatesV)
        
        # Determine whether we need to estimate the model
        estimate = evaluate_em_convergence(freq_estimation, CurrDates, PrevDates, t)
        if R_new is None:
            estimate = 1
        
        if estimate == 1:
            print(f"Now I am estimating the model: {dt.strftime('%d-%m-%Y')}")
            from .kalman import reset_kf_run_flag
            reset_kf_run_flag()
            slice_end = min(int(iQ + 10), nobs)
            R_new = fit_dfm_em(X_new[:slice_end, :], P, None, run_id=t+1, debug_dir=debug_dir)
            R_new['Groups'] = Group
            R_new['Series'] = Series
            
            beta, XX = ar_realtime_gdp(X_new, i_ser_idx, p_ar, iQ, native=native, debug_dir=debug_dir)
            print("Now I am updating the prediction\n")

        # Prediction
        if t > vintages_to_run[0]:
            y_old = y_new.copy()
            
        y_new, v_miss, X_Last, Date = update_predictions(
            X_old, X_new, X_Last, R_new, Date, iQ, i_ser_idx, nobs, debug_dir=debug_dir
        )
        
        if t == vintages_to_run[0]:
            y_old = y_new.copy()
            
        temp1, temp2 = compute_annual_nowcast(y_new, X_new, X_Last, CurrDates, iQ, i_ser_idx, debug_dir=debug_dir)
        t1, t2, t3, t4, _ = update_benchmark_predictions(X_new, beta, XX, p_ar, iQ, i_ser_idx, debug_dir=debug_dir)
        fe_m = compute_forecast_errors(fe_m, v_miss, ReleaseTime, t, y_new, i_ser_idx, X_Last, iQ, Date, t1, t2, t3, debug_dir=debug_dir)
        current_output = nowcast_current_output(current_output, ReleaseTime, y_new, Date, v_miss, X_Last, t, i_ser_idx, iQ, debug_dir=debug_dir)
        
        # Store predictions
        AnnualPrediction[t, 0] = temp1 
        AnnualGDP[t, 0] = temp2        
        AR_Back[t, 0] = t1
        AR_Now[t, 0] = t2
        AR_Fore[t, 0] = t3
        RW[t, 0] = t4
        
        Backcast[t, 0]   = y_new[0, i_ser_idx]
        Nowcast[t, 0]    = y_new[1, i_ser_idx]
        Forecast[t, 0]   = y_new[2, i_ser_idx]
        Forecast2S[t, 0] = y_new[3, i_ser_idx]
        Forecast3S[t, 0] = y_new[4, i_ser_idx]
        
        target_indices = np.array([iQ-3, iQ, iQ+3, iQ+6, iQ+9], dtype=int)
        Actual[t, :] = X_Last[target_indices, i_ser_idx]
        if Date.ndim > 1:
            Reference[t, :] = Date[target_indices, 0]
        else:
            Reference[t, :] = Date[target_indices]

        # Always run para_constdg and compute standard deviations
        slice_end = min(int(iQ + 10), nobs)
        from .kalman import para_constdg
        from .prediction import (
            estimate_breach_probability,
            estimate_annual_breach_probability,
            estimate_sequential_breach_probability
        )
        res_smooth = para_constdg(X_new[:slice_end, :], R_new, 0, debug_dir=debug_dir)
        
        direction = threshold_direction if threshold_direction is not None else 'lower'
        thresh_val = threshold if threshold is not None else 0.0
        
        p_back, _, std_back = estimate_breach_probability(res_smooth, R_new, i_ser_idx, iQ - 3, thresh_val, X_new=X_new[:slice_end, :], dist=prob_distribution, direction=direction)
        p_now, _, std_now  = estimate_breach_probability(res_smooth, R_new, i_ser_idx, iQ, thresh_val, X_new=X_new[:slice_end, :], dist=prob_distribution, direction=direction)
        p_fore, _, std_fore = estimate_breach_probability(res_smooth, R_new, i_ser_idx, iQ + 3, thresh_val, X_new=X_new[:slice_end, :], dist=prob_distribution, direction=direction)
        p_f2s, _, std_f2s  = estimate_breach_probability(res_smooth, R_new, i_ser_idx, iQ + 6, thresh_val, X_new=X_new[:slice_end, :], dist=prob_distribution, direction=direction)
        p_f3s, _, std_f3s  = estimate_breach_probability(res_smooth, R_new, i_ser_idx, iQ + 9, thresh_val, X_new=X_new[:slice_end, :], dist=prob_distribution, direction=direction)
        
        try:
            p_ann, _, std_ann, _, _ = estimate_annual_breach_probability(
                res_smooth, R_new, i_ser_idx, dt.year, Date, data_file, thresh_val, X_new=X_new[:slice_end, :], dist=prob_distribution, direction=direction
            )
        except Exception as e:
            print(f"Warning: could not calculate annual probability for vintage {t}: {e}")
            p_ann = np.nan
            std_ann = np.nan
            
        try:
            p_seq, _, _ = estimate_sequential_breach_probability(
                res_smooth, R_new, i_ser_idx, int(iQ), Date,
                threshold=seq_threshold,
                direction=seq_direction,
                n_consecutive=seq_n_consecutive,
                horizon=seq_horizon,
                k_offset=seq_k_offset
            )
        except Exception as e:
            print(f"Warning: could not calculate sequential breach probability for vintage {t}: {e}")
            p_seq = np.nan
            
        Prob_Seq[t, 0] = p_seq
        Std_Backcast[t, 0]    = std_back
        Std_Nowcast[t, 0]     = std_now
        Std_Forecast[t, 0]    = std_fore
        Std_Forecast2S[t, 0]  = std_f2s
        Std_Forecast3S[t, 0]  = std_f3s
        Std_Annual[t, 0]      = std_ann

        if threshold is not None:
            Prob_Backcast[t, 0]   = p_back
            Prob_Nowcast[t, 0]    = p_now
            Prob_Forecast[t, 0]   = p_fore
            Prob_Forecast2S[t, 0] = p_f2s
            Prob_Forecast3S[t, 0] = p_f3s
            Prob_Annual[t, 0]     = p_ann

    elapsed_time = time.time() - start_time
    print(f"Historical Evaluation Complete! Time elapsed: {elapsed_time:.2f} seconds")

    # Export Final Outputs
    os.makedirs(output_dir, exist_ok=True)

    def csvwrite(filename, data):
        filepath = os.path.join(output_dir, filename)
        if isinstance(data, (pd.DataFrame, pd.Series)):
            data = data.to_numpy()
        data = np.array(data, dtype=float)
        if data.ndim == 0: data = data.reshape(1, 1)
        elif data.ndim == 1: data = data.reshape(-1, 1)
        np.savetxt(filepath, data, delimiter=",", fmt="%.16g")
        with open(filepath, "r") as f: content = f.read()
        with open(filepath, "w") as f: f.write(content.replace("nan", "NaN"))

    csvwrite("AnnualPrediction.csv", AnnualPrediction)
    csvwrite("AnnualGDP.csv", AnnualGDP)
    csvwrite("Backcast.csv", Backcast)
    csvwrite("Nowcast.csv", Nowcast)
    csvwrite("Forecast.csv", Forecast)
    csvwrite("Forecast2S.csv", Forecast2S)
    csvwrite("Forecast3S.csv", Forecast3S)
    csvwrite("Actual.csv", Actual)
    csvwrite("Reference.csv", Reference)
    csvwrite("RW.csv", RW)
    csvwrite("AR_Back.csv", AR_Back)
    csvwrite("AR_Now.csv", AR_Now)
    csvwrite("AR_Fore.csv", AR_Fore)
    csvwrite("ReleaseTime.csv", ReleaseTime)

    csvwrite("Std_Backcast.csv", Std_Backcast)
    csvwrite("Std_Nowcast.csv", Std_Nowcast)
    csvwrite("Std_Forecast.csv", Std_Forecast)
    csvwrite("Std_Forecast2S.csv", Std_Forecast2S)
    csvwrite("Std_Forecast3S.csv", Std_Forecast3S)
    csvwrite("Std_Annual.csv", Std_Annual)
    csvwrite("Prob_Seq.csv", Prob_Seq)

    if threshold is not None:
        csvwrite("Prob_Backcast.csv", Prob_Backcast)
        csvwrite("Prob_Nowcast.csv", Prob_Nowcast)
        csvwrite("Prob_Forecast.csv", Prob_Forecast)
        csvwrite("Prob_Forecast2S.csv", Prob_Forecast2S)
        csvwrite("Prob_Forecast3S.csv", Prob_Forecast3S)
        csvwrite("Prob_Annual.csv", Prob_Annual)

    print(f"All outputs exported successfully to: {output_dir}")

    # Export Excel Report (identical to MATLAB SaveResultToExcel.m)
    try:
        import datetime
        timestamp_str = datetime.datetime.now().strftime("%m%d%y_%H%M")
        excel_path = os.path.join(output_dir, f"Result_{timestamp_str}.xlsx")
        from openpyxl import Workbook
        
        wb = Workbook()
        
        # 1. Sheet: Prediction
        ws_pred = wb.active
        ws_pred.title = 'Prediction'
        
        # Slice arrays to only include vintages that were run
        run_indices = list(vintages_to_run)
        n_run = len(run_indices)
        
        Reference_run = Reference[run_indices, :]
        ReleaseTime_run = ReleaseTime[run_indices, :]
        Actual_run = Actual[run_indices, :]
        Backcast_run = Backcast[run_indices, :]
        Nowcast_run = Nowcast[run_indices, :]
        Forecast_run = Forecast[run_indices, :]
        Forecast2S_run = Forecast2S[run_indices, :]
        Forecast3S_run = Forecast3S[run_indices, :]
        AnnualPrediction_run = AnnualPrediction[run_indices, :]
        
        if threshold is not None:
            Prob_Backcast_run = Prob_Backcast[run_indices, :]
            Std_Backcast_run = Std_Backcast[run_indices, :]
            Prob_Nowcast_run = Prob_Nowcast[run_indices, :]
            Std_Nowcast_run = Std_Nowcast[run_indices, :]
            Prob_Forecast_run = Prob_Forecast[run_indices, :]
            Std_Forecast_run = Std_Forecast[run_indices, :]
            Prob_Forecast2S_run = Prob_Forecast2S[run_indices, :]
            Std_Forecast2S_run = Std_Forecast2S[run_indices, :]
            Prob_Forecast3S_run = Prob_Forecast3S[run_indices, :]
            Std_Forecast3S_run = Std_Forecast3S[run_indices, :]
            Prob_Annual_run = Prob_Annual[run_indices, :]
            Std_Annual_run = Std_Annual[run_indices, :]
        
        ref_dates = pd.to_datetime(Reference_run[:, 1] - 693960, unit='D', origin='1899-12-30').date
        
        # Safely convert ReleaseTime_run to floats
        rel_time_floats = np.array([float(x[0]) if isinstance(x, (list, np.ndarray)) else float(x) for x in ReleaseTime_run.flatten()])
        vin_dates = pd.to_datetime(rel_time_floats - 693960, unit='D', origin='1899-12-30').date
        
        pred_headers = ['Reference Quarter', 'Day Prediction', 'Actual', 'Backcast', 'Nowcast', 'Forecast', '2-step', '3-step', 'Annual Nowcast']
        if threshold is not None:
            sign = '<' if threshold_direction.lower() == 'lower' else '>'
            pred_headers += [
                f'Prob Backcast {sign} {threshold}', 'Std Backcast',
                f'Prob Nowcast {sign} {threshold}', 'Std Nowcast',
                f'Prob Forecast {sign} {threshold}', 'Std Forecast',
                f'Prob 2-step {sign} {threshold}', 'Std 2-step',
                f'Prob 3-step {sign} {threshold}', 'Std 3-step',
                f'Prob Annual Nowcast {sign} {threshold}', 'Std Annual'
            ]
        ws_pred.append(pred_headers)
        pred_rows = []
        
        for idx in range(n_run):
            row_data = [
                ref_dates[idx],
                vin_dates[idx],
                Actual_run[idx, 1],
                Backcast_run[idx, 0],
                Nowcast_run[idx, 0],
                Forecast_run[idx, 0],
                Forecast2S_run[idx, 0],
                Forecast3S_run[idx, 0],
                AnnualPrediction_run[idx, 0]
            ]
            if threshold is not None:
                row_data += [
                    Prob_Backcast_run[idx, 0], Std_Backcast_run[idx, 0],
                    Prob_Nowcast_run[idx, 0], Std_Nowcast_run[idx, 0],
                    Prob_Forecast_run[idx, 0], Std_Forecast_run[idx, 0],
                    Prob_Forecast2S_run[idx, 0], Std_Forecast2S_run[idx, 0],
                    Prob_Forecast3S_run[idx, 0], Std_Forecast3S_run[idx, 0],
                    Prob_Annual_run[idx, 0], Std_Annual_run[idx, 0]
                ]
            ws_pred.append(row_data)
            pred_rows.append(row_data)
            
        # 2. Sheet: Database
        ws_db = wb.create_sheet('Database')
        db_headers = ['Date'] + Series
        ws_db.append(db_headers)
        
        excel_dates = Date.flatten() - 693960
        for idx in range(len(excel_dates)):
            row_data = [excel_dates[idx]] + list(X_Last[idx, :])
            ws_db.append(row_data)
            
        # 3. Sheet: Indicators
        ws_ind = wb.create_sheet('Indicators')
        ws_ind['A1'] = 'Indicator'
        ws_ind['B1'] = 'Relase day'
        
        for idx, ser in enumerate(Series):
            ws_ind.cell(row=idx+2, column=1, value=ser)
        for idx, row in enumerate(ReleaseDay):
            val1 = float(row[0]) if not pd.isna(row[0]) else None
            val2 = float(row[1]) if not pd.isna(row[1]) else None
            ws_ind.cell(row=idx+2, column=2, value=val1)
            ws_ind.cell(row=idx+2, column=3, value=val2)
        for idx, q_ind in enumerate(SeriesQ):
            ws_ind.cell(row=idx+1, column=5, value=q_ind)
            
        # 4. Sheets: Q2-2026 Target, Q3-2026 Target, Q4-2026 Target (Target Horizon Monitoring)
        b0, b1 = prob_bins if prob_bins is not None else (5.0, 5.35)
        
        q_targets = {
            'Q2-2026': {
                '3-step': (pd.Timestamp('2025-07-01'), pd.Timestamp('2025-09-30')),
                '2-step': (pd.Timestamp('2025-10-01'), pd.Timestamp('2025-12-31')),
                'Forecast': (pd.Timestamp('2026-01-01'), pd.Timestamp('2026-03-31')),
                'Nowcast': (pd.Timestamp('2026-04-01'), pd.Timestamp('2026-06-30')),
                'Backcast': (pd.Timestamp('2026-07-01'), pd.Timestamp('2026-08-31')),
            },
            'Q3-2026': {
                '3-step': (pd.Timestamp('2025-10-01'), pd.Timestamp('2025-12-31')),
                '2-step': (pd.Timestamp('2026-01-01'), pd.Timestamp('2026-03-31')),
                'Forecast': (pd.Timestamp('2026-04-01'), pd.Timestamp('2026-06-30')),
                'Nowcast': (pd.Timestamp('2026-07-01'), pd.Timestamp('2026-09-30')),
                'Backcast': (pd.Timestamp('2026-10-01'), pd.Timestamp('2026-11-30')),
            },
            'Q4-2026': {
                '3-step': (pd.Timestamp('2026-01-01'), pd.Timestamp('2026-03-31')),
                '2-step': (pd.Timestamp('2026-04-01'), pd.Timestamp('2026-06-30')),
                'Forecast': (pd.Timestamp('2026-07-01'), pd.Timestamp('2026-09-30')),
                'Nowcast': (pd.Timestamp('2026-10-01'), pd.Timestamp('2026-12-31')),
                'Backcast': (pd.Timestamp('2027-01-01'), pd.Timestamp('2027-02-28')),
            }
        }

        ind_release_map = {
            1: 'PMI EM', 2: 'PMI EM', 5: 'PDB', 6: 'Ekspor', 7: 'Impor Barang Konsumsi',
            8: 'Impor Barang Modal', 9: 'Impor Bahan Baku dan Barang Penolong',
            15: 'Penjualan Mobil', 17: 'PMI EM', 23: 'M2 (Uang Beredar)', 27: 'Konsumsi Semen'
        }

        from scipy.stats import norm
        for q_code, stages in q_targets.items():
            ws_q = wb.create_sheet(f'{q_code} Target')
            q_headers = [
                'Day Prediction', 'Target Horizon', 'Growth Estimate (%)', 'Nowcasting FY 2026 (%)', 
                'Delta Target (% pt)', 'Indikator Pemicu', 'Arah Dampak', 'Phase', 'Std Error (%)', 
                f'P(< {b0}%)', f'P({b0}% - {b1}%)', f'P(> {b1}%)'
            ]
            ws_q.append(q_headers)
            
            rows_target = []
            for col_name, (start_dt, end_dt) in stages.items():
                for idx in range(n_run):
                    vin_dt = pd.to_datetime(vin_dates[idx])
                    if start_dt <= vin_dt <= end_dt:
                        if col_name == 'Backcast':
                            mu_val = float(Backcast_run[idx, 0])
                            phase_lbl = 'Backcast'
                        elif col_name == 'Nowcast':
                            mu_val = float(Nowcast_run[idx, 0])
                            phase_lbl = 'Nowcast'
                        elif col_name == 'Forecast':
                            mu_val = float(Forecast_run[idx, 0])
                            phase_lbl = '1-step Forecast'
                        elif col_name == '2-step':
                            mu_val = float(Forecast2_run[idx, 0]) if 'Forecast2_run' in locals() else float(Forecast_run[idx, 0])
                            phase_lbl = '2-step Forecast'
                        else:
                            mu_val = float(Forecast3_run[idx, 0]) if 'Forecast3_run' in locals() else float(Forecast_run[idx, 0])
                            phase_lbl = '3-step Forecast'
                            
                        fy_val = float(AnnualNowcast_run[idx, 0]) if 'AnnualNowcast_run' in locals() else mu_val
                        std_val = float(Std_Nowcast_run[idx, 0]) if threshold is not None else 1.0
                        
                        day_num = vin_dt.day
                        if day_num in ind_release_map:
                            ind_driver = ind_release_map[day_num]
                        elif day_num in [18, 19, 20]:
                            ind_driver = 'Suku Bunga'
                        else:
                            closest = min(ind_release_map.keys(), key=lambda k: abs(k - day_num))
                            ind_driver = ind_release_map[closest] if abs(closest - day_num) <= 2 else 'Indikator Bulanan'
                            
                        rows_target.append({
                            'Day Prediction': vin_dates[idx],
                            'Target Horizon': q_code,
                            'Growth Estimate (%)': round(mu_val, 4),
                            'Nowcasting FY 2026 (%)': round(fy_val, 4),
                            'Indikator Pemicu': ind_driver,
                            'Phase': phase_lbl,
                            'Std Error (%)': round(std_val, 4)
                        })
                        
            df_t = pd.DataFrame(rows_target)
            if not df_t.empty:
                df_t['Delta Target (% pt)'] = df_t['Growth Estimate (%)'].diff().round(4).fillna(0.0)
                def get_impact(d):
                    if d > 0.005: return 'Peningkatan (Upward)'
                    elif d < -0.005: return 'Penurunan (Downward)'
                    else: return 'Tetap (No Change)'
                df_t['Arah Dampak'] = df_t['Delta Target (% pt)'].apply(get_impact)
                
                mu_arr = df_t['Growth Estimate (%)'].values
                std_arr = df_t['Std Error (%)'].values
                df_t[f'P(< {b0}%)'] = np.round(norm.cdf(b0, loc=mu_arr, scale=std_arr) * 100.0, 2)
                df_t[f'P({b0}% - {b1}%)'] = np.round((norm.cdf(b1, loc=mu_arr, scale=std_arr) - norm.cdf(b0, loc=mu_arr, scale=std_arr)) * 100.0, 2)
                df_t[f'P(> {b1}%)'] = np.round((1.0 - norm.cdf(b1, loc=mu_arr, scale=std_arr)) * 100.0, 2)
                
                ordered_cols = [
                    'Day Prediction', 'Target Horizon', 'Growth Estimate (%)', 'Nowcasting FY 2026 (%)',
                    'Delta Target (% pt)', 'Indikator Pemicu', 'Arah Dampak', 'Phase', 'Std Error (%)',
                    f'P(< {b0}%)', f'P({b0}% - {b1}%)', f'P(> {b1}%)'
                ]
                df_t = df_t[ordered_cols]
                
                for row_vals in df_t.values.tolist():
                    ws_q.append(row_vals)
                    
            # Save standalone CSV for each target quarter
            csv_q_path = os.path.join(output_dir, f'{q_code.replace("-", "_")}_Forecast_and_Probabilities.csv')
            df_t.to_csv(csv_q_path, index=False)
            print(f"Exported {q_code} Target CSV dataset to: {csv_q_path}")

        wb.save(excel_path)
        print(f"Excel report successfully exported to: {excel_path}")
        
        # Copy to named file matching MATLAB pattern
        named_excel_path = os.path.join(output_dir, f"Nowcasting report_{timestamp_str}.xlsx")
        shutil.copyfile(excel_path, named_excel_path)
        print(f"Copied Excel report to: {named_excel_path}")
        
        # Generate Native PowerPoint Presentations if requested
        if generate_pptx:
            try:
                from .pptx_exporter import generate_probability_pptx
                effective_bins = prob_bins if prob_bins is not None else (5.0, 5.35)
                df_pred_export = pd.DataFrame(pred_rows, columns=pred_headers)
                generate_probability_pptx(
                    df_pred_export,
                    std_annual=Std_Annual_run if threshold is not None else None,
                    std_nowcast=Std_Nowcast_run if threshold is not None else None,
                    bins=effective_bins,
                    output_dir=output_dir
                )
            except Exception as e_pptx:
                print(f"Warning: Could not generate PowerPoint report: {e_pptx}")
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"Warning: Failed to generate Excel report: {e}")


# Legacy aliases for backward compatibility
ml_build_calendar = build_calendar
ml_pseudo_real_time_vintages = build_pseudo_real_time_vintages
ml_current_quarter = get_current_quarter
ml_update_prediction = update_predictions
ml_annual_nowcast = compute_annual_nowcast
ml_nowcast_current_output = nowcast_current_output
em_dfm_ss_block_idioqarma_restrmq = fit_dfm_em
ml_mixed_frequency_restrictions = mixed_frequency_restrictions
ml_law_motion_idiosyncratic = idiosyncratic_law_of_motion
ml_estimation_yn = evaluate_em_convergence
ml_ar_realtime_gdp = ar_realtime_gdp
ml_update_benchmark_predictions = update_benchmark_predictions
ml_nowcast_forecast_error = compute_forecast_errors

