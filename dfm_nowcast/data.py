import os
import warnings
import datetime
import numpy as np
import pandas as pd
from scipy.interpolate import CubicSpline
from scipy.signal import lfilter

from .utils import (
    init_debug_dir,
    csvwrite_debug,
    to_matlab_datenum,
    to_excel_serial,
    get_matlab_xlsread_matrix,
    matlab_datevec,
    convert_date
)

def verify_dates(df, name):
    """
    Checks for duplicate dates or non-standard date formats in the first column of the DataFrame.
    """
    if df.empty or df.shape[1] == 0:
        return
    dates_col = df.iloc[:, 0]
    
    # 1. Check duplicate dates
    duplicates = dates_col[dates_col.duplicated()]
    if not duplicates.empty:
        warnings.warn(f"Duplicate date entries found in {name} date column: {duplicates.tolist()}", RuntimeWarning)
        
    # 2. Check non-standard date formats
    non_standard = []
    for val in dates_col.dropna():
        if isinstance(val, (int, float, np.integer, np.floating)):
            continue
        if isinstance(val, (pd.Timestamp, datetime.date, datetime.datetime)):
            continue
        try:
            pd.to_datetime(val)
        except Exception:
            non_standard.append(val)
            
    if non_standard:
        warnings.warn(f"Non-standard date formats found in {name} date column: {non_standard[:10]}", RuntimeWarning)


def verify_data_sheets(data_dict):
    """
    Verifies that all required sheets and columns are present and well-formed.
    """
    required_keys = ['InfoM', 'MonthlyData', 'InfoQ', 'QuarterlyData', 'Calendar']
    for key in required_keys:
        if key not in data_dict or data_dict[key] is None or data_dict[key].empty:
            raise ValueError(f"Malformed input data: key '{key}' is missing, empty, or None.")
            
    # Check InfoM columns
    required_info_m = ['INCLUDED', 'Indicator Code', 'log', 'MoM', 'YoY', 'Months lag', 'Days lag']
    for col in required_info_m:
        if col not in data_dict['InfoM'].columns:
            raise ValueError(f"Malformed input data: 'InfoM' sheet is missing required column '{col}'.")
            
    # Check InfoQ columns
    required_info_q = ['INCLUDED', 'Indicator Code', 'log', 'QoQ', 'YoY', 'Months lag', 'Days lag']
    for col in required_info_q:
        if col not in data_dict['InfoQ'].columns:
            raise ValueError(f"Malformed input data: 'InfoQ' sheet is missing required column '{col}'.")
            
    # Verify included indicators exist in Data sheets
    info_m = data_dict['InfoM']
    included_m = info_m.loc[info_m['INCLUDED'] == 1, 'Indicator Code'].tolist()
    for col in included_m:
        if col not in data_dict['MonthlyData'].columns:
            raise ValueError(f"Malformed input data: Indicator '{col}' is marked as included in InfoM but is missing from MonthlyData.")
            
    info_q = data_dict['InfoQ']
    included_q = info_q.loc[info_q['INCLUDED'] == 1, 'Indicator Code'].tolist()
    for col in included_q:
        if col not in data_dict['QuarterlyData'].columns:
            raise ValueError(f"Malformed input data: Indicator '{col}' is marked as included in InfoQ but is missing from QuarterlyData.")


def check_collinearity(data_dict, transformed_data=None, series=None, series_q=None):
    """
    Diagnoses collinearity in both the raw and transformed/aligned datasets.
    Prints diagnostic output and raises warnings if high collinearity is detected.
    """
    # 1. Raw Monthly Data
    if 'MonthlyData' in data_dict and 'InfoM' in data_dict:
        df_m = data_dict['MonthlyData']
        info_m = data_dict['InfoM']
        included_m = info_m['INCLUDED'].values == 1
        cols_m = info_m.loc[included_m, 'Indicator Code'].tolist()
        if all(c in df_m.columns for c in cols_m) and cols_m:
            df_m_clean = df_m[cols_m]
            _analyze_collinearity_df(df_m_clean, "Raw Monthly Variables")
            
    # 2. Raw Quarterly Data
    if 'QuarterlyData' in data_dict and 'InfoQ' in data_dict:
        df_q = data_dict['QuarterlyData']
        info_q = data_dict['InfoQ']
        included_q = info_q['INCLUDED'].values == 1
        cols_q = info_q.loc[included_q, 'Indicator Code'].tolist()
        if all(c in df_q.columns for c in cols_q) and cols_q:
            df_q_clean = df_q[cols_q]
            _analyze_collinearity_df(df_q_clean, "Raw Quarterly Variables")
            
    # 3. Transformed aligned data
    if transformed_data is not None and series is not None:
        df_trans = pd.DataFrame(transformed_data, columns=series)
        if series_q is not None:
            series_m = [s for s in series if s not in series_q]
            if series_m:
                _analyze_collinearity_df(df_trans[series_m], "Transformed Monthly Variables")
            _analyze_collinearity_df(df_trans[series_q], "Transformed Quarterly Variables")
        else:
            _analyze_collinearity_df(df_trans, "Transformed Variables")


def _analyze_collinearity_df(df, label):
    # Calculate variances
    variances = df.var()
    constant_cols = variances[variances == 0].index.tolist()
    if constant_cols:
        warnings.warn(f"[{label}] Constant columns (variance=0) found: {constant_cols}", RuntimeWarning)
        
    corr_matrix = df.corr()
    valid_cols = corr_matrix.columns[~corr_matrix.isna().all()]
    if len(valid_cols) < 2:
        return
        
    sub_corr = corr_matrix.loc[valid_cols, valid_cols].fillna(0.0)
    
    try:
        eigvals = np.linalg.eigvalsh(sub_corr)
        min_eig = np.min(eigvals)
        max_eig = np.max(eigvals)
        condition_number = max_eig / max(min_eig, 1e-15)
        
        if condition_number > 1e4:
            warnings.warn(f"[{label}] Extremely Ill-conditioned/Multicollinear matrix! Condition Number: {condition_number:.4e}", RuntimeWarning)
    except Exception:
        pass
        
    # Check for highly correlated pairs
    high_corr_pairs = []
    columns = corr_matrix.columns
    for i in range(len(columns)):
        for j in range(i + 1, len(columns)):
            r = corr_matrix.iloc[i, j]
            if not np.isnan(r) and abs(r) >= 0.90:
                high_corr_pairs.append((columns[i], columns[j], r))
                
    if len(high_corr_pairs) > 0:
        high_corr_pairs.sort(key=lambda x: abs(x[2]), reverse=True)
        pair_strings = [f"{v1} & {v2} (r={r:.3f})" for v1, v2, r in high_corr_pairs[:3]]
        warnings.warn(f"[{label}] Highly correlated pairs found: {', '.join(pair_strings)}", RuntimeWarning)

    clean_df = df.dropna(how='all', axis=1)
    clean_df = clean_df.loc[:, clean_df.var() > 1e-8]
    clean_df_imputed = clean_df.fillna(clean_df.mean())
    
    if clean_df_imputed.shape[1] > 1:
        try:
            standardized_data = (clean_df_imputed - clean_df_imputed.mean()) / clean_df_imputed.std()
            U, S, Vt = np.linalg.svd(standardized_data, full_matrices=False)
            
            smallest_sv_idx = np.argmin(S)
            if S[smallest_sv_idx] < 1.0:
                contrib = np.abs(Vt[smallest_sv_idx, :])
                top_contrib_indices = np.argsort(contrib)[::-1][:3]
                contributors = [standardized_data.columns[idx] for idx in top_contrib_indices]
                warnings.warn(f"[{label}] Smallest singular value is {S[smallest_sv_idx]:.4f}. Top contributors to collinearity direction: {contributors}", RuntimeWarning)
        except Exception:
            pass


def load_data_input(data_input):
    """
    Normalizes data_input into a dictionary of pandas DataFrames.
    Keys: 'InfoM', 'MonthlyData', 'InfoQ', 'QuarterlyData', 'Calendar'
    """
    import json
    if isinstance(data_input, dict):
        normalized = {}
        for key in ['InfoM', 'MonthlyData', 'InfoQ', 'QuarterlyData', 'Calendar']:
            val = data_input.get(key)
            if val is None:
                raise ValueError(f"Missing required key '{key}' in data_input dict.")
            if isinstance(val, pd.DataFrame):
                normalized[key] = val.copy()
            elif isinstance(val, str) and os.path.exists(val):
                normalized[key] = pd.read_csv(val)
            else:
                normalized[key] = pd.DataFrame(val)
    elif isinstance(data_input, str):
        if not os.path.exists(data_input):
            raise FileNotFoundError(f"Input path '{data_input}' does not exist.")
            
        if data_input.endswith('.xlsx'):
            xls = pd.ExcelFile(data_input)
            normalized = {
                'InfoM': pd.read_excel(xls, 'InfoM'),
                'MonthlyData': pd.read_excel(xls, 'MonthlyData'),
                'InfoQ': pd.read_excel(xls, 'InfoQ'),
                'QuarterlyData': pd.read_excel(xls, 'QuarterlyData'),
                'Calendar': pd.read_excel(xls, 'Calendar', header=None)
            }
            
        elif data_input.endswith('.json'):
            with open(data_input, 'r', encoding='utf-8') as f:
                data_dict = json.load(f)
            normalized = {}
            for key in ['InfoM', 'MonthlyData', 'InfoQ', 'QuarterlyData', 'Calendar']:
                val = data_dict.get(key)
                if val is None:
                    raise ValueError(f"Missing required key '{key}' in JSON data.")
                normalized[key] = pd.DataFrame(val)
            
        elif os.path.isdir(data_input):
            normalized = {}
            for key in ['InfoM', 'MonthlyData', 'InfoQ', 'QuarterlyData', 'Calendar']:
                csv_path = os.path.join(data_input, f"{key}.csv")
                if not os.path.exists(csv_path):
                    csv_path = os.path.join(data_input, f"{key.lower()}.csv")
                if not os.path.exists(csv_path):
                    raise FileNotFoundError(f"Missing required CSV file '{key}.csv' in directory '{data_input}'.")
                if key == 'Calendar':
                    normalized[key] = pd.read_csv(csv_path, header=None)
                else:
                    normalized[key] = pd.read_csv(csv_path)
            
        else:
            raise ValueError(f"Unsupported file format or path: {data_input}")
    else:
        raise TypeError(f"data_input must be a path string or dict, got {type(data_input)}")

    verify_data_sheets(normalized)
    verify_dates(normalized['MonthlyData'], 'MonthlyData')
    verify_dates(normalized['QuarterlyData'], 'QuarterlyData')
    return normalized


def read_data(data_input, start_est=None, date_format='matlab', debug_dir=None):
    """
    Replicates the MATLAB ReadData function perfectly mirroring MATLAB file I/O quirks.
    Supports Excel path, JSON path, CSV folder path, or dict of DataFrames.
    """
    warnings.filterwarnings("ignore", category=RuntimeWarning)

    # Environment Setup for debug exports
    target_dir = init_debug_dir('ReadData', debug_dir=debug_dir)

    data_dict = load_data_input(data_input)

    # ==== STEP 1 = Load and Manipulate monthly data ====
    info_m = data_dict['InfoM']
    
    included_m = info_m['INCLUDED'].values == 1
    list_m = np.where(included_m)[0] + 1
    
    series_m = info_m.loc[included_m, 'Indicator Code'].tolist()
    transf_m = info_m.loc[included_m, ['log', 'MoM', 'YoY']].values
    release_day_m = info_m.loc[included_m, ['Months lag', 'Days lag']].values

    monthly_data = data_dict['MonthlyData']
    
    dates_mv_excel = monthly_data.iloc[:, 0]
    dates_mv = convert_date(dates_mv_excel, to_format='matlab')
    t_len = len(dates_mv)

    data_m = monthly_data[series_m].values

    csvwrite_debug(target_dir, 'ListM.csv', list_m)
    csvwrite_debug(target_dir, 'TransfM.csv', transf_m)
    csvwrite_debug(target_dir, 'ReleaseDayM.csv', release_day_m)
    csvwrite_debug(target_dir, 'DatesMV.csv', dates_mv)
    csvwrite_debug(target_dir, 'DataM.csv', data_m)
    csvwrite_debug(target_dir, 'SeriesM.csv', series_m)

    # Transformations
    data_mm = data_m.copy().astype(float)
    
    j1 = (transf_m[:, 0] == 1)
    data_mm[:, j1] = 100 * np.log(data_mm[:, j1])

    j2 = (transf_m[:, 1] == 1)
    data_mm[1:, j2] = data_mm[1:, j2] - data_mm[:-1, j2]
    
    j3 = (transf_m[:, 2] == 1)
    data_mm[12:, j3] = data_mm[12:, j3] - data_mm[:-12, j3]

    data_mm[0, j2] = np.nan
    data_mm[:12, j3] = np.nan

    data_m_trf = data_mm.copy()
    
    csvwrite_debug(target_dir, 'J1.csv', j1)
    csvwrite_debug(target_dir, 'J2.csv', j2)
    csvwrite_debug(target_dir, 'J3.csv', j3)
    csvwrite_debug(target_dir, 'DataMM.csv', data_mm)
    csvwrite_debug(target_dir, 'DataMTrf.csv', data_m_trf)

    # ==== STEP 2 = Load and Manipulate quarterly data ====
    info_q = data_dict['InfoQ']
    
    included_q = info_q['INCLUDED'].values == 1
    list_q = np.where(included_q)[0] + 1
    
    series_q = info_q.loc[included_q, 'Indicator Code'].tolist()
    transf_q = info_q.loc[included_q, ['log', 'QoQ', 'YoY']].values
    release_day_q = info_q.loc[included_q, ['Months lag', 'Days lag']].values

    quarterly_data = data_dict['QuarterlyData']
    
    dates_qv_excel = quarterly_data.iloc[:, 0]
    
    if not pd.api.types.is_numeric_dtype(dates_qv_excel) and not pd.api.types.is_datetime64_any_dtype(dates_qv_excel):
        dates_qv_dt = pd.to_datetime(dates_qv_excel)
    elif pd.api.types.is_datetime64_any_dtype(dates_qv_excel):
        dates_qv_dt = dates_qv_excel
    else:
        dates_qv_dt = pd.to_datetime(dates_qv_excel, unit='D', origin='1899-12-30')
        
    dates_qv = np.column_stack((dates_qv_dt.dt.year, dates_qv_dt.dt.month))
    
    data_q = quarterly_data[series_q].values

    csvwrite_debug(target_dir, 'ListQ.csv', list_q)
    csvwrite_debug(target_dir, 'TransfQ.csv', transf_q)
    csvwrite_debug(target_dir, 'ReleaseDayQ.csv', release_day_q)
    csvwrite_debug(target_dir, 'DatesQV.csv', dates_qv)
    csvwrite_debug(target_dir, 'DataQ.csv', data_q)
    csvwrite_debug(target_dir, 'SeriesQ.csv', series_q)

    # Transformations
    data_qq = data_q.copy().astype(float)
    
    j1_q = (transf_q[:, 0] == 1)
    data_qq[:, j1_q] = 100 * np.log(data_qq[:, j1_q])
    
    j2_q = (transf_q[:, 1] == 1)
    data_qq[1:, j2_q] = data_qq[1:, j2_q] - data_qq[:-1, j2_q]
    
    j3_q = (transf_q[:, 2] == 1)
    data_qq[4:, j3_q] = data_qq[4:, j3_q] - data_qq[:-4, j3_q]
    
    data_qq[0, j2_q] = np.nan
    data_qq[:4, j3_q] = np.nan

    data_q_trf = data_qq.copy()
    
    kron_multiplier = np.array([[np.nan], [np.nan], [1]])
    data_qm_trf = np.kron(data_q_trf, kron_multiplier)
    
    t_q = data_qm_trf.shape[0]
    n_q = data_qm_trf.shape[1]
    if t_len > t_q:
        data_qm_trf = np.vstack([data_qm_trf, np.full((t_len - t_q, n_q), np.nan)])

    csvwrite_debug(target_dir, 'DataQQ.csv', data_qq)
    csvwrite_debug(target_dir, 'DataQTrf.csv', data_q_trf)
    csvwrite_debug(target_dir, 'DataQMTrf.csv', data_qm_trf)

    # ==== STEP 3 = Merge Monthly and Quarterly ====
    data = np.hstack([data_m_trf, data_qm_trf])
    series = series_m + series_q
    release_day = np.vstack([release_day_m, release_day_q])

    monthly_dates_col = monthly_data.iloc[:, 0]
    if not pd.api.types.is_numeric_dtype(monthly_dates_col) and not pd.api.types.is_datetime64_any_dtype(monthly_dates_col):
         dates_mv_dt = pd.to_datetime(monthly_dates_col)
    elif pd.api.types.is_datetime64_any_dtype(monthly_dates_col):
         dates_mv_dt = monthly_dates_col
    else:
         dates_mv_dt = pd.to_datetime(monthly_dates_col, unit='D', origin='1899-12-30')
         
    tt = np.column_stack([
        dates_mv_dt.dt.year, dates_mv_dt.dt.month, dates_mv_dt.dt.day,
        dates_mv_dt.dt.hour, dates_mv_dt.dt.minute, dates_mv_dt.dt.second
    ])
    
    if start_est is None:
        non_nan_indices = np.where(~np.isnan(data_m).any(axis=1))[0]
        if len(non_nan_indices) > 0:
            first_idx = non_nan_indices[0]
            start_est = [int(tt[first_idx, 0]), int(tt[first_idx, 1])]
        else:
            start_est = [int(tt[0, 0]), int(tt[0, 1])]
            
    i_est_arr = np.where((tt[:, 0] == start_est[0]) & (tt[:, 1] == start_est[1]))[0]
    i_est = i_est_arr[0] if len(i_est_arr) > 0 else 0
        
    data = data[i_est:, :]
    dates = dates_mv[i_est:]

    csvwrite_debug(target_dir, 'TT.csv', tt)
    csvwrite_debug(target_dir, 'iEst.csv', i_est + 1) 

    csvwrite_debug(target_dir, 'Data.csv', data)
    csvwrite_debug(target_dir, 'ReleaseDay.csv', release_day)
    csvwrite_debug(target_dir, 'Dates.csv', dates)
    csvwrite_debug(target_dir, 'nQ.csv', n_q)
    csvwrite_debug(target_dir, 'Series.csv', series)

    if date_format != 'matlab':
        dates = convert_date(dates, to_format=date_format, from_format='matlab')

    check_collinearity(data_dict, data, series, series_q)

    return data, series, release_day, series_q, dates, n_q


def build_pseudo_real_time_vintages(x_last, date_array, name_vintage, mm, r, pmv, debug_dir=None):
    """
    Replicates the MATLAB ML_PseudoRealTimeVintages function.
    Constructs pseudo real-time data vintages based on the release calendar.
    """
    warnings.filterwarnings("ignore", category=RuntimeWarning)

    target_dir = init_debug_dir('ML_PseudoRealTimeVintages', debug_dir=debug_dir)
    
    # Initialize the vintage matrix with NaNs
    v = np.full(x_last.shape, np.nan)
    
    # Extract [Year, Month] from the dates
    temp = matlab_datevec(date_array)
    dates_v = temp[:, 0:2]

    csvwrite_debug(target_dir, 'DatesV.csv', dates_v)
    csvwrite_debug(target_dir, 'V_initial.csv', v)

    num_t = len(name_vintage)
    num_i = mm.shape[0]
    
    cm_trace = np.full((num_t, num_i), np.nan)
    idx_trace = np.full((num_t, num_i), np.nan)

    dd = []
    dd_name = []

    for t in range(num_t):
        for i in range(num_i):
            valid_r_indices = np.where(r[i, :] <= name_vintage[t])[0]
            
            if len(valid_r_indices) > 0:
                cm = np.max(valid_r_indices)
                cm_trace[t, i] = cm + 1
                
                idx_arr = np.where((dates_v[:, 0] == pmv[cm, 0]) & (dates_v[:, 1] == pmv[cm, 1]))[0]
                
                if len(idx_arr) > 0:
                    idx = idx_arr[0]
                    idx_trace[t, i] = idx + 1
                    v[0:idx+1, i] = x_last[0:idx+1, i]
                    
        dd.append(v.copy())
        dd_name.append(name_vintage[t])
        
        csvwrite_debug(target_dir, f'DD_t{t+1}.csv', v)
        csvwrite_debug(target_dir, f'DDname_t{t+1}.csv', name_vintage[t])

    csvwrite_debug(target_dir, 'CM_trace.csv', cm_trace)
    csvwrite_debug(target_dir, 'IDX_trace.csv', idx_trace)

    return dd, dd_name, v, dates_v


# Legacy alias
ml_pseudo_real_time_vintages = build_pseudo_real_time_vintages


def build_calendar(data_input, start_eval=None, last_eval=None, n_q=None, hist_eval=True, date_format='matlab', debug_dir=None):
    """
    Replicates the MATLAB ML_BuildCalendar function perfectly.
    """
    warnings.filterwarnings("ignore", category=RuntimeWarning)

    import datetime
    current_year = datetime.datetime.now().year
    last_year = current_year - 1
    if hist_eval:
        if start_eval is None:
            start_eval = f"1-Jan-{last_year}"
        if last_eval is None:
            last_eval = f"31-Dec-{last_year}"
    else:
        if start_eval is None:
            start_eval = f"1-Jan-{current_year}"
        if last_eval is None:
            last_eval = f"31-Dec-{current_year}"

    if n_q is None:
        data_dict = load_data_input(data_input)
        info_q = data_dict['InfoQ']
        n_q = int(np.sum(info_q['INCLUDED'].values == 1))

    target_dir = init_debug_dir('ML_BuildCalendar', debug_dir=debug_dir)

    a = get_matlab_xlsread_matrix(data_input, 'Calendar')
        
    # Check column 1 (INCLUDE flags) or column 0 for included rows
    col1 = a[1:-1, 1] if a.shape[1] > 1 else a[1:-1, 0]
    try:
        col1_num = pd.to_numeric(col1, errors='coerce')
        list_idx = np.where(col1_num > 0)[0]
    except Exception:
        list_idx = np.array([])
    if len(list_idx) == 0:
        list_idx = np.arange(len(a[1:-1]))
    list_matlab = list_idx + 1

    csvwrite_debug(target_dir, 'List.csv', list_matlab)

    # ==== STEP 1 = Release Data Calendar Matrix ====
    mm_raw = a[1:-1, 1:]
    mm_raw = mm_raw[list_idx, :]
    
    mm = np.vstack([mm_raw[n_q:, :], mm_raw[:n_q, :]])
    
    j = np.isnan(mm)
    r = mm.copy()
    
    r[~j] = mm[~j] + 693960
    
    r_flat = r.flatten()
    r_valid = np.unique(r_flat[~np.isnan(r_flat)])
    r_nans = r_flat[np.isnan(r_flat)]
    l_unique = np.concatenate([r_valid, r_nans])

    csvwrite_debug(target_dir, 'MM.csv', mm)
    csvwrite_debug(target_dir, 'J_matrix.csv', j.astype(float))
    csvwrite_debug(target_dir, 'R_dates.csv', r)
    csvwrite_debug(target_dir, 'L_unique.csv', l_unique)

    # ==== STEP 2 = Period Processing ====
    p = a[0, 1:].reshape(-1, 1)  
    j_p = np.isnan(p)
    p_matlab = p.copy()
    
    p_matlab[~j_p] = p[~j_p] + 693960
    
    period = matlab_datevec(p_matlab)
    pmv = period[:, [0, 1]]  

    csvwrite_debug(target_dir, 'p_vector.csv', p_matlab)
    csvwrite_debug(target_dir, 'Period.csv', period)
    csvwrite_debug(target_dir, 'PMV.csv', pmv)

    # ==== STEP 3 = Start / End Evaluation Identification ====
    start_eval_dn = pd.to_datetime(start_eval).toordinal() + 366
    last_eval_dn = pd.to_datetime(last_eval).toordinal() + 366

    start_idx = np.sum(l_unique <= start_eval_dn) 
    
    if start_idx > 0 and start_idx < len(l_unique):
        year_start = pd.to_datetime(l_unique[start_idx - 1] - 693960, unit='D', origin='1899-12-30').year
        year_next = pd.to_datetime(l_unique[start_idx] - 693960, unit='D', origin='1899-12-30').year
        if year_start != year_next:
            start_idx += 1
            
    last_idx = np.sum(l_unique <= last_eval_dn)

    csvwrite_debug(target_dir, 'start_idx.csv', start_idx)
    csvwrite_debug(target_dir, 'last_idx.csv', last_idx)

    l_sliced = l_unique[(start_idx - 1) : last_idx]
    name_vintage = l_sliced

    csvwrite_debug(target_dir, 'name_vintage.csv', name_vintage)

    if date_format != 'matlab':
        name_vintage = convert_date(name_vintage, to_format=date_format, from_format='matlab')
        r = convert_date(r, to_format=date_format, from_format='matlab')

    return name_vintage, pmv, mm, r


# Legacy alias
ml_build_calendar = build_calendar


def interpolate_missing_spline(X, options, debug_dir=None):
    """
    Replicates the MATLAB remNaNs / remNaNs_spline function.
    Handles missing values via trimming, median fills, moving averages, and cubic splines.
    """
    warnings.filterwarnings("ignore", category=RuntimeWarning)

    target_dir = init_debug_dir('remNaNs_spline', debug_dir=debug_dir)
    
    # Work on a copy so we don't accidentally mutate the original array
    X = np.array(X, copy=True, dtype=float)
    T, N = X.shape
    k = int(options.get('k', 3))
    method = int(options.get('method', 1))
    
    indNaN = np.isnan(X)
    
    # Helper function for Moving Average fill
    def apply_ma_fill(x_col, is_nan_mask):
        median_val = np.nanmedian(x_col)
        x_col[is_nan_mask] = median_val
        
        b = np.ones(2 * k + 1) / (2 * k + 1)
        padded_x = np.concatenate((np.full(k, x_col[0]), x_col, np.full(k, x_col[-1])))
        x_MA = lfilter(b, 1, padded_x)
        x_MA = x_MA[2 * k :]
        
        x_col[is_nan_mask] = x_MA[is_nan_mask]
        return x_col

    # Helper function for Spline fill
    def apply_spline_and_ma(x_col):
        isnanx = np.isnan(x_col)
        valid_idx = np.where(~isnanx)[0]
        
        if len(valid_idx) > 1:
            t1 = valid_idx.min()
            t2 = valid_idx.max()
            
            cs = CubicSpline(valid_idx, x_col[valid_idx])
            x_col[t1 : t2 + 1] = cs(np.arange(t1, t2 + 1))
            
        isnanx = np.isnan(x_col)
        x_col = apply_ma_fill(x_col, isnanx)
        return x_col

    if method == 1:
        for i in range(N):
            X[:, i] = apply_ma_fill(X[:, i], indNaN[:, i])
            
    elif method in [2, 3, 4]:
        if method == 2:
            rem1 = np.sum(indNaN, axis=1) > (N * 0.8)
        else:
            rem1 = np.sum(indNaN, axis=1) == N
            
        t_arr = np.arange(1, T + 1)
        nanLead = np.cumsum(rem1) == t_arr
        
        nanEnd = np.cumsum(rem1[::-1]) == t_arr
        nanEnd = nanEnd[::-1]
        
        nanLE = nanLead | nanEnd
        
        X = X[~nanLE, :]
        indNaN = np.isnan(X)
        
        if method in [2, 4]:
            for i in range(N):
                X[:, i] = apply_spline_and_ma(X[:, i])
                
    elif method == 5:
        for i in range(N):
            X[:, i] = apply_spline_and_ma(X[:, i])

    csvwrite_debug(target_dir, f'X_out_method{method}.csv', X)
    csvwrite_debug(target_dir, f'indNaN_out_method{method}.csv', indNaN.astype(float))

    return X, indNaN


# Legacy alias
remnans_spline = interpolate_missing_spline


class DFMPreprocessor:
    """
    Robust scaling and winsorization preprocessing pipeline for Dynamic Factor Models.
    Matches standard scikit-learn API with fit, transform, fit_transform, and inverse_transform.
    """
    def __init__(self, winsorization=False, winsorization_k=4.0, robust_scaling=False):
        self.winsorization = winsorization
        self.winsorization_k = winsorization_k
        self.robust_scaling = robust_scaling
        self.Mx = None
        self.Wx = None

    def fit(self, X, y=None):
        X_wins = X.copy()
        if self.winsorization:
            T, N = X_wins.shape
            for j in range(N):
                col = X_wins[:, j]
                med = np.nanmedian(col)
                mad = np.nanmedian(np.abs(col - med))
                if mad < 1e-6:
                    mad = 1.0
                lower_bound = med - self.winsorization_k * mad
                upper_bound = med + self.winsorization_k * mad
                non_nan_mask = ~np.isnan(col)
                X_wins[non_nan_mask, j] = np.clip(col[non_nan_mask], lower_bound, upper_bound)

        if self.robust_scaling:
            self.Mx = np.nanmedian(X_wins, axis=0).reshape(1, -1)
            q75, q25 = np.nanpercentile(X_wins, [75, 25], axis=0)
            self.Wx = (q75 - q25).reshape(1, -1)
            self.Wx[self.Wx < 1e-12] = 1.0
        else:
            self.Mx = np.nanmean(X_wins, axis=0).reshape(1, -1)
            self.Wx = np.nanstd(X_wins, axis=0, ddof=1).reshape(1, -1)
            self.Wx[self.Wx < 1e-12] = 1.0
        return self

    def transform(self, X):
        if self.Mx is None or self.Wx is None:
            raise ValueError("DFMPreprocessor must be fitted before transforming data.")
        X_wins = X.copy()
        if self.winsorization:
            T, N = X_wins.shape
            for j in range(N):
                col = X_wins[:, j]
                med = np.nanmedian(col)
                mad = np.nanmedian(np.abs(col - med))
                if mad < 1e-6:
                    mad = 1.0
                lower_bound = med - self.winsorization_k * mad
                upper_bound = med + self.winsorization_k * mad
                non_nan_mask = ~np.isnan(col)
                X_wins[non_nan_mask, j] = np.clip(col[non_nan_mask], lower_bound, upper_bound)

        T_len = X_wins.shape[0]
        return (X_wins - np.tile(self.Mx, (T_len, 1))) / np.tile(self.Wx, (T_len, 1))

    def fit_transform(self, X, y=None):
        return self.fit(X).transform(X)

    def inverse_transform(self, X_scaled):
        if self.Mx is None or self.Wx is None:
            raise ValueError("DFMPreprocessor must be fitted before inverse-transforming data.")
        T_len = X_scaled.shape[0]
        return X_scaled * np.tile(self.Wx, (T_len, 1)) + np.tile(self.Mx, (T_len, 1))


def update_dataset_from_zaki(target_file='data/INO_130726.xlsx', zaki_file='data/Data Zaki.xlsx', output_file=None):
    """
    Updates the target DFM Excel dataset (e.g. INO_130726.xlsx) with latest indicator data
    from Zaki's Excel export (Data Zaki.xlsx).

    Preserves Excel formula structures, formatting, and sheet metadata while ensuring
    Calendar and Data sheets are cleanly evaluated for DFM execution.
    """
    import openpyxl

    if output_file is None:
        output_file = target_file

    if not os.path.exists(target_file):
        raise FileNotFoundError(f"Target Excel file not found: {target_file}")
    if not os.path.exists(zaki_file):
        raise FileNotFoundError(f"Zaki Excel file not found: {zaki_file}")

    # Load evaluated values for matching date serials & header names
    wb_val = openpyxl.load_workbook(target_file, data_only=True)

    # Load un-evaluated workbook for editing to preserve formulas
    wb_raw = openpyxl.load_workbook(target_file, data_only=False)

    # Read Zaki data sheets
    zaki_m = pd.read_excel(zaki_file, sheet_name='Monthly')
    zaki_q = pd.read_excel(zaki_file, sheet_name='Quarterly')

    zaki_m['Date'] = pd.to_datetime(zaki_m.iloc[:, 0])
    zaki_q['Date'] = pd.to_datetime(zaki_q.iloc[:, 0])

    # Series mapping dictionary (INO Code -> Zaki Column Name)
    default_map_m = {
        'Man_PMI_EM': 'Manufacturing PMI: Headline: sa: Emerging Markets',
        'M2': 'Broad Money (M2)',
        'Car_sales': 'Motor Vehicle Sales: PT Astra: Local',
        'Cement': 'Consumption: Cement: Domestic',
        'Export': 'Export: fob',
        'M_con_goods': 'Imports: Consumer Goods (CG)',
        'M_cap_goods': 'Imports: Capital Goods',
        'M_raw_mat': 'Imports: Intermediate Goods (IG)',
        'Inflation': 'Headline Inflation',
        'Forex': 'Spot FX Rate: Month-End: Bank Indonesia: IDR to US Dollar',
        'Policy_rate': 'Policy Rate: Month End: Indonesia: BI-Rate'
    }

    default_map_q = {
        'RGDP_level': 'Gross Domestic Product: SNA 2008: 2010p',
        'RGDP_growth': 'Gross Domestic Product: SNA 2008: 2010p: YoY %',
        'Bus_Act': 'Business Survey: Business Activity: R: Weighted Net Balance '
    }

    log_updates = []

    # Update MonthlyData
    if 'MonthlyData' in wb_raw.sheetnames:
        ws_m_val = wb_val['MonthlyData']
        ws_m_raw = wb_raw['MonthlyData']

        m_headers = {ws_m_val.cell(row=1, column=c).value: c for c in range(1, ws_m_val.max_column + 1) if ws_m_val.cell(row=1, column=c).value is not None}
        
        # Ensure row 1 explicit headers & Column A date serials are set in raw sheet
        for h_name, col_idx in m_headers.items():
            ws_m_raw.cell(row=1, column=col_idx).value = str(h_name)

        m_date_to_row = {}
        for r in range(2, ws_m_val.max_row + 1):
            val = ws_m_val.cell(row=r, column=1).value
            if val is not None:
                ws_m_raw.cell(row=r, column=1).value = val
                dt = convert_date(val)
                if pd.notna(dt):
                    m_date_to_row[dt.strftime('%Y-%m-%d')] = r

        for code, zcol in default_map_m.items():
            if code in m_headers and zcol in zaki_m.columns:
                col_idx = m_headers[code]
                z_series = zaki_m[['Date', zcol]].dropna()
                for _, row in z_series.iterrows():
                    d_str = row['Date'].strftime('%Y-%m-%d')
                    val = float(row[zcol])
                    if d_str in m_date_to_row:
                        r_idx = m_date_to_row[d_str]
                        old_val = ws_m_val.cell(row=r_idx, column=col_idx).value
                        if old_val is None or pd.isna(old_val) or not np.isclose(float(old_val), val, rtol=1e-5):
                            ws_m_raw.cell(row=r_idx, column=col_idx).value = val
                            log_updates.append({
                                'Sheet': 'MonthlyData',
                                'Indicator': code,
                                'Date': d_str,
                                'OldValue': old_val,
                                'NewValue': val
                            })

    # Update QuarterlyData
    if 'QuarterlyData' in wb_raw.sheetnames:
        ws_q_val = wb_val['QuarterlyData']
        ws_q_raw = wb_raw['QuarterlyData']

        q_headers = {ws_q_val.cell(row=1, column=c).value: c for c in range(1, ws_q_val.max_column + 1) if ws_q_val.cell(row=1, column=c).value is not None}
        
        # Ensure row 1 explicit headers & Column A date serials are set in raw sheet
        for h_name, col_idx in q_headers.items():
            ws_q_raw.cell(row=1, column=col_idx).value = str(h_name)

        q_date_to_row = {}
        for r in range(2, ws_q_val.max_row + 1):
            val = ws_q_val.cell(row=r, column=1).value
            if val is not None:
                ws_q_raw.cell(row=r, column=1).value = val
                dt = convert_date(val)
                if pd.notna(dt):
                    q_date_to_row[dt.strftime('%Y-%m-%d')] = r

        for code, zcol in default_map_q.items():
            if code in q_headers and zcol in zaki_q.columns:
                col_idx = q_headers[code]
                z_series = zaki_q[['Date', zcol]].dropna()
                for _, row in z_series.iterrows():
                    d_str = row['Date'].strftime('%Y-%m-%d')
                    val = float(row[zcol])
                    if d_str in q_date_to_row:
                        r_idx = q_date_to_row[d_str]
                        old_val = ws_q_val.cell(row=r_idx, column=col_idx).value
                        if old_val is None or pd.isna(old_val) or not np.isclose(float(old_val), val, rtol=1e-5):
                            ws_q_raw.cell(row=r_idx, column=col_idx).value = val
                            log_updates.append({
                                'Sheet': 'QuarterlyData',
                                'Indicator': code,
                                'Date': d_str,
                                'OldValue': old_val,
                                'NewValue': val
                            })

    # Evaluate & Populate Calendar sheet
    info_q = pd.read_excel(target_file, sheet_name='InfoQ')
    info_m = pd.read_excel(target_file, sheet_name='InfoM')
    monthly_df = pd.read_excel(target_file, sheet_name='MonthlyData')

    m_dates = [convert_date(x) for x in monthly_df['Indicator Code'] if pd.notna(x)]
    date_serials = [to_excel_serial(d) for d in m_dates]

    cal_rows = []
    # Quarterly indicators
    for _, row in info_q.iterrows():
        code = row['Indicator Code']
        inc = row['INCLUDED']
        m_lag = int(row['Months lag'])
        d_lag = int(row['Days lag'])
        
        r_vals = [code, inc]
        for d in m_dates:
            if d.month in [3, 6, 9, 12]:
                pub_year = d.year + (d.month + m_lag - 1) // 12
                pub_month = (d.month + m_lag - 1) % 12 + 1
                pub_dt = pd.Timestamp(year=pub_year, month=pub_month, day=d_lag)
                r_vals.append(to_excel_serial(pub_dt))
            else:
                r_vals.append(None)
        cal_rows.append(r_vals)

    # Monthly indicators
    for _, row in info_m.iterrows():
        code = row['Indicator Code']
        inc = row['INCLUDED']
        m_lag = int(row['Months lag'])
        d_lag = int(row['Days lag'])
        
        r_vals = [code, inc]
        for d in m_dates:
            pub_year = d.year + (d.month + m_lag - 1) // 12
            pub_month = (d.month + m_lag - 1) % 12 + 1
            pub_dt = pd.Timestamp(year=pub_year, month=pub_month, day=d_lag)
            r_vals.append(to_excel_serial(pub_dt))
        cal_rows.append(r_vals)

    header_row = ['Economic Indicators', 'INCLUDE'] + date_serials

    if 'Calendar' in wb_raw.sheetnames:
        del wb_raw['Calendar']
    ws_cal = wb_raw.create_sheet('Calendar')
    ws_cal.append(header_row)
    for r in cal_rows:
        ws_cal.append(r)

    wb_raw.save(output_file)
    wb_val.close()
    wb_raw.close()

    # Ensure Octave io package compatibility by fixing absolute Target paths in xl/_rels/workbook.xml.rels
    try:
        import zipfile
        with zipfile.ZipFile(output_file, 'r') as zin:
            file_data = {item.filename: zin.read(item.filename) for item in zin.infolist()}
        
        rels_key = 'xl/_rels/workbook.xml.rels'
        if rels_key in file_data:
            text = file_data[rels_key].decode('utf-8')
            if 'Target="/xl/worksheets/' in text:
                text = text.replace('Target="/xl/worksheets/', 'Target="worksheets/')
                file_data[rels_key] = text.encode('utf-8')
                
                with zipfile.ZipFile(output_file, 'w') as zout:
                    for name, content in file_data.items():
                        zout.writestr(name, content)
    except Exception:
        pass

    return log_updates




